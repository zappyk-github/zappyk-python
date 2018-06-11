#!/usr/bin/env python3
# -*- coding: utf-8 -*-
__author__ = 'zappyk'

import sys, argparse #, configparser
import csv
import PIL
import xlrd #, xlwt
import openpyxl

from xlutils import copy

from lib_zappyk._os_file import _basename, _fileExist, _copy2
from lib_zappyk._string  import _trim
from lib_zappyk._log     import _log

_version = '0.1'

_project = 'XLSx2RewriteOut'

_description = '''
Rewrite a Spreadsheet ( .xls / .xlsx ) on source CSV file.
'''

_epilog = "Version: %s" % _version

_default_csv_delimiter = ';'
_default_csv_quotechar = None

logs = _log(); logs.setFormat(None);

###############################################################################
def _getargs():
#CZ#parser = argparse.ArgumentParser(description=_description, epilog=_epilog, formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    formatter = lambda prog: argparse.HelpFormatter(prog, max_help_position=50, width=120)
    formatter = lambda prog: argparse.ArgumentDefaultsHelpFormatter(prog, max_help_position=50, width=120)
#CZ#formatter = lambda prog: argparse.RawDescriptionHelpFormatter(prog, max_help_position=50, width=120)
    parser = argparse.ArgumentParser(description=_description, epilog=_epilog, formatter_class=formatter) #, argument_default=not argparse.SUPPRESS)

#CZ#pgroup.add_argument('-p' , '--power'        , help='display a power of a given number'   , type=int           , choices=[1,2,3,4,5])
#CZ#pgroup.add_argument('-s' , '--square'       , help='display a square of a given number'  , type=int)
    parser.add_argument('-d' , '--debug'        , help='increase output debug'               , action='count'     , default=0)
    parser.add_argument('-v' , '--verbose'      , help='output verbosity'                    , action='store_true')
    parser.add_argument('-V' , '--version'      , help='print version number'                , action='version'   , version='%(prog)s '+_version)
    parser.add_argument('-fs', '--file_source'  , help='file source CSV for rewrite layout'  , type=str           , required=True)
    parser.add_argument('-fl', '--file_layout'  , help='file spreadsheet (only .xls/.xlsx)'  , type=str           , required=True)
    parser.add_argument('-fo', '--file_output'  , help='file output merge on source-layout'  , type=str           , required=True)
    parser.add_argument('-cv', '--cell_values'  , help='modify cell value format: "1,1=test"', type=str           , nargs='+')
    parser.add_argument('-ci', '--cell_images'  , help='insert cell image format: "1,1=file"', type=str           , nargs='+')
    parser.add_argument('-cs', '--csv_delimiter', help='csv delimiter char'                  , type=str           , default=_default_csv_delimiter)
    parser.add_argument('-cd', '--csv_quotechar', help='csv quote char'                      , type=str           , default=_default_csv_quotechar)
#CZ#parser.add_argument('name'                  , help='Name')
#CZ#parser.add_argument('surname'               , help='Surname')

    args = parser.parse_args()

    return(args)

###############################################################################
class ExceptionFileInputNotParsed(Exception):
    def __init__(self, value):
        self.value = value
    def __str__(self):
        return repr(self.value)

###############################################################################
def open_file_input(file_input):
    workbook = None
    xlsx_ext = None

    exception = None

    if workbook is None:
        try:
            workbook = open_file_xlsx(file_input)
            xlsx_ext = True
        except Exception as e:
            exception = e
            pass
        finally:
            if args.debug:
                logs.info('...try open .xlsx format... ', '')
                if workbook is None:
                    logs.info('no! :-(')
                else:
                    logs.info('yes :-)')

    if workbook is None:
        try:
            workbook = open_file_xls_(file_input)
            xlsx_ext = False
        except Exception as e:
            exception = e
            pass
        finally:
            if args.debug:
                logs.info('...try open .xls  format... ', '')
                if workbook is None:
                    logs.info('no! :-(')
                else:
                    logs.info('yes :-)')

    if workbook is None:
        if exception is not None:
            if str(exception) == 'formatting_info=True not yet implemented':
            #CZ#raise ExceptionFileInputNotParsed('File is not .xls/.xlsx format!')
                raise Exception('File is not .xls/.xlsx format!')
            else:
                raise Exception(str(exception))
        else:
            raise Exception('Something is wrong, file not opened!')
    else:
        logs.info('Opened file successfully')

    return(workbook, xlsx_ext)

###############################################################################
def open_file_xlsx(file_input):
#CZ#workbook = openpyxl.load_workbook(file_input, data_only=True)
    workbook = openpyxl.load_workbook(file_input)
    return(workbook)

###############################################################################
def open_file_xls_(file_input):
#CZ#workbook = xlrd.open_workbook(file_input)
    workbook = xlrd.open_workbook(file_input, formatting_info=True)
    return(workbook)

###############################################################################
def split_cell_story(string):
    if args.debug:
        log.info("#=coordinate=> [ %s ]" % (string))

    char_split_1 = '='
    char_split_2 = ','
    string_coordinate = string.split(char_split_1)[0]
    string_cell_story = string.split(char_split_1)[1:]
    coordinate = string_coordinate.split(char_split_2)
    cell_story = char_split_1.join(string_cell_story)

    coordinate_XY = None
    if len(coordinate) == 1:
        from openpyxl.utils import coordinate_from_string, column_index_from_string
        x_y = coordinate_from_string(coordinate[0])      # returns ('A',4)
        col = column_index_from_string(x_y[0])  # returns 1
        row = x_y[1]
        coordinate_XY = coordinate[0]
        coordinate[0] = row
        coordinate.append(col)

    if args.debug:
        log.info("#=coordinate=> [ %s | %s ]=[%s]" % (coordinate[0], coordinate[1], cell_story))

    return(coordinate[0], coordinate[1], coordinate_XY, cell_story)

###############################################################################
if __name__ == '__main__':
    args = _getargs()

    file_source   = args.file_source
    file_layout   = args.file_layout
    file_output   = args.file_output
    cell_values   = args.cell_values
    cell_images   = args.cell_images
    csv_delimiter = args.csv_delimiter
    csv_quotechar = args.csv_quotechar

    name_source = _basename(file_source)
    name_layout = _basename(file_layout)
    name_output = _basename(file_output)

    open_layout = file_layout

    try:
        logs.info('Rewrite name layout %s from %s file and marge on %s file output' % (name_layout, name_source, name_output))

    #CZ#logs.info('Copy file layout "%s" for layout in "%s" file output:' % (file_layout, file_output))
    #CZ#_copy2(file_layout, file_output); open_layout = file_output

        logs.info('Opened file layout "%s" for rewrite:' % open_layout)
        (workbook
        ,xlsx_ext) = open_file_input(open_layout)

        readcsv = []
        with open(file_source, 'r') as fs:
            linecsv = csv.reader(fs, delimiter=csv_delimiter, quotechar=csv_quotechar)
            readcsv = list(linecsv)
#CZ#except ExceptionFileInputNotParsed as e:
#CZ#    logs.error(str(e))
    except Exception as e:
        logs.error(str(e))

    workbook_output = None
    workbook_outxls = None
    if xlsx_ext:
    #CZ#workbook_output = open_file_xlsx(file_output)
    #CZ#workbook_output = open_file_xlsx(open_layout)
        workbook_output = workbook
    else:
    #CZ#logs.info('Copy file layout "%s" for layout in "%s" file output:' % (file_layout, file_output))
    #CZ#_copy2(file_layout, file_output); open_layout = file_output

    #CZ#workbook_output = open_file_xls_(file_output)
        workbook_output = open_file_xls_(open_layout)
    #CZ#workbook_output = workbook
    #CZ#workbook_outxls = copy.copy(workbook_output)

    try:
        logs.info(' * Set worksheet 0...')
        if xlsx_ext:
            workbook_output.active = 0
            worksheet = workbook_output.active
        else:
            worksheet = workbook_output.sheet_by_index(0)

        logs.info(' * Modify values...')
        if xlsx_ext:
            #______________
            # fill   CSV  :
            ####################################################################
            row = 0
            while row < len(readcsv):
                if args.debug:
                    logs.info("line row %3s. = %s" % (row, readcsv[row]))
                col = 0
                while col < len(readcsv[row]):
                    val = readcsv[row][col]
                    if _trim(val) != '':
                        worksheet.cell(row=row+1, column=col+1).value = val
                        if args.debug >= 2:
                            logs.info("write [r.%3s|c.%3s]=[%s]" % (row, col, val))
                    col += 1
                row += 1
            #______________
            # cell values :
            ####################################################################
            if cell_values is not None:
                for cell_value in cell_values:
                    (cell_row
                    ,cell_col
                    ,cell_x_y
                    ,cell_val)= split_cell_story(cell_value)
                    worksheet.cell(row=cell_row, column=cell_col).value = cell_val
            #worksheet['J2']  = 999
            #worksheet['J9']  = '!!! PROVA !!!'
            #worksheet['J46'] = 99.99
            #worksheet.cell(row=  2, column= 10).value = 99
            #worksheet.cell(row=  9, column= 10).value = '!!! PROVA !!!'
            #worksheet.cell(row= 46, column= 10).value = 99.99
            #______________
            # cell images :
            ####################################################################
            if cell_images is not None:
                for cell_image in cell_images:
                    (cell_row
                    ,cell_col
                    ,cell_x_y
                    ,cell_img)= split_cell_story(cell_image)
                    if _fileExist(cell_img):
                        logs.info(' * Insert image %s ...' % _basename(cell_img))
                        load_img = openpyxl.drawing.image.Image(cell_img)
                    #CZ#load_img.anchor(worksheet.cell('A1')); worksheet.add_image(load_img)
                    #CZ#worksheet.add_image(load_img, 'A1')
                        worksheet.add_image(load_img, cell_x_y)
                    else:
                        logs.info(' * Insert image %s NOT FOUND!' % _basename(cell_img))
        else:
            #______________
            # fill   CSV  :
            ####################################################################
            #______________
            # cell values :
            ####################################################################
            if cell_values is not None:
                for cell_value in cell_values:
                    (cell_row
                    ,cell_col
                    ,cell_x_y
                    ,cell_val)= split_cell_story(cell_value)
                    worksheet.cell(cell_row, cell_col).value = cell_val
            #worksheet.cell(      2,         10).value = 999
            #worksheet.cell(      9,         10).value = '!!! PROVA !!!'
            #worksheet.cell(     46,         10).value = 99.99
            #______________
            # cell images :
            ####################################################################

        logs.info(' * Write output!')
        if xlsx_ext:
            workbook_output.save(file_output)
        else:
            workbook_outxls = copy.copy(workbook_output)
            workbook_outxls.save(file_output)

    except Exception as e:
    #CZ#logs.info('Something is wrong when delete (-- Sheets[···]) or (Write Out [···]) file :-|')
    #CZ#if xlsx_ext:
    #CZ#    logs.warning('[%s]' % workbook_output.get_sheet_names())
    #CZ#else:
    #CZ#    logs.warning('[%s]' % [worksheet.name for worksheet in workbook_outxls._Workbook__worksheets])
        logs.error(str(e))

    sys.exit(0)
